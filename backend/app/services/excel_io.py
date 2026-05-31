"""Excel import and export."""

from __future__ import annotations

import io
import re
import unicodedata
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database import fetch_all, get_company_by_slug, grade_columns_list, dedupe_price_rows
from app.services.normalizer import normalize_device_name
from app.services.summary_builder import build_summary

MODEL_HEADERS = frozenset(
    {
        "דגם",
        "model",
        "device",
        "devices",
        "מכשיר",
        "שם דגם",
        "device name",
        "model name",
    }
)

# Dynamica spreadsheets often use grade letter suffix/prefix order (תקין A vs A תקין).
GRADE_HEADER_ALIASES: dict[str, dict[str, list[str]]] = {
    "dynamica": {
        "a": ["A תקין", "תקין A", "A"],
        "b": ["B תקין", "תקין B", "B"],
        "c": [
            "C שבור / סדוק",
            "C שבור/סדוק",
            "C - שבור / סדוק",
            "C - שבור/סדוק",
            "שבור / סדוק C",
            "שבור/סדוק C",
            "שבור / סדוק",
            "שבור/סדוק",
            "C",
        ],
        "d": ["D תקול", "תקול D", "D"],
    },
    "partner": {
        "a": ["תקין"],
    },
}

HEADER_SCAN_ROWS = 15


def _load_aligned_rows(ws, *, max_row: Optional[int] = None) -> list[tuple[Any, ...]]:
    """Read rows with fixed columns 1..max_column so indices match across all rows."""
    last_row = max_row if max_row is not None else (ws.max_row or HEADER_SCAN_ROWS)
    last_col = ws.max_column or 1
    scan_rows = max(last_row, HEADER_SCAN_ROWS)
    return list(
        ws.iter_rows(
            min_row=1,
            max_row=scan_rows,
            min_col=1,
            max_col=last_col,
            values_only=True,
        )
    )


def _header_leading_blanks(headers: list[str]) -> int:
    for i, header in enumerate(headers):
        if _normalize_header(header):
            return i
    return 0


def _data_col_idx(row: tuple[Any, ...], header_idx: int, leading_blanks: int) -> int:
    """Map header column index to data row when data starts further left than headers."""
    if leading_blanks <= 0 or header_idx < 0:
        return header_idx
    if row and row[0] not in (None, ""):
        return max(header_idx - leading_blanks, 0)
    return header_idx


def _cell_looks_like_device(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    if _parse_price(text) is not None and not any(ch.isalpha() for ch in text):
        return False
    return normalize_device_name(text) is not None


def _resolve_row_columns(
    row: tuple[Any, ...],
    model_idx: int,
    grade_idx: int,
    leading_blanks: int,
) -> tuple[int, int]:
    """Pick device/price columns by cell content, not only header index."""
    header_model = _data_col_idx(row, model_idx, leading_blanks)
    header_grade = _data_col_idx(row, grade_idx, leading_blanks)
    candidates: list[tuple[int, int, int]] = []

    indices = {
        model_idx,
        grade_idx,
        header_model,
        header_grade,
        model_idx - leading_blanks,
        grade_idx - leading_blanks,
    }
    for mc in indices:
        for gc in indices:
            if mc < 0 or gc < 0 or mc == gc:
                continue
            if mc >= len(row) or gc >= len(row):
                continue
            if not _cell_looks_like_device(row[mc]):
                continue
            if _parse_price(row[gc]) is None:
                continue
            score = 0
            if gc == grade_idx:
                score += 20
            if gc == header_grade:
                score += 15
            if abs(gc - grade_idx) == 1:
                score += 5
            if mc == model_idx:
                score += 10
            if mc == header_model:
                score += 8
            if abs(mc - model_idx) == 1:
                score += 3
            candidates.append((score, mc, gc))

    if candidates:
        candidates.sort(reverse=True)
        _, mc, gc = candidates[0]
        return mc, gc

    return header_model, header_grade


def _parse_price(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace("₪", "").replace(",", "").strip()
        if not text or text in {"-", "—", "–"}:
            return None
        value = text
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


DYNAMICA_GRADE_RES: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"(?:^|\s)A(?:\s|$)|תקין\s*A|A\s*תקין", re.IGNORECASE), "a"),
    (re.compile(r"(?:^|\s)B(?:\s|$)|תקין\s*B|B\s*תקין", re.IGNORECASE), "b"),
    (re.compile(r"שבור|סדוק|(?:^|\s)C(?:\s|$)", re.IGNORECASE), "c"),
    (re.compile(r"תקול|(?:^|\s)D(?:\s|$)", re.IGNORECASE), "d"),
)


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\ufeff", "")
    text = re.sub(r"[\u200e\u200f\u202a-\u202e]", "", text)
    text = re.sub(r"[\u2010-\u2015\-–—‑]", " ", text)
    text = re.sub(r"\s*/\s*", " / ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _is_model_header(header: str) -> bool:
    normalized = _normalize_header(header).lower()
    if not normalized:
        return False
    if normalized in MODEL_HEADERS:
        return True
    if "דגם" in header:
        return True
    return normalized.endswith(" model") or normalized.endswith(" device")


def _build_grade_label_map(slug: str, grades: list[dict]) -> dict[str, str]:
    label_to_key: dict[str, str] = {}
    for grade in grades:
        key = grade["key"]
        labels = {grade["label"], grade["label"].replace(" / ", "/"), grade["label"].replace("/", " / ")}
        labels.update(GRADE_HEADER_ALIASES.get(slug, {}).get(key, []))
        for label in labels:
            label_to_key[_normalize_header(label)] = key
            label_to_key[_normalize_header(label).lower()] = key
    return label_to_key


def _match_dynamica_grade(header: str) -> Optional[str]:
    normalized = _normalize_header(header)
    if not normalized:
        return None
    for pattern, key in DYNAMICA_GRADE_RES:
        if pattern.search(normalized):
            return key
    return None


def _match_grade_column(header: str, label_to_key: dict[str, str], *, slug: Optional[str] = None) -> Optional[str]:
    normalized = _normalize_header(header)
    if not normalized:
        return None
    if normalized in label_to_key:
        return label_to_key[normalized]
    lower = normalized.lower()
    if lower in label_to_key:
        return label_to_key[lower]
    if slug == "partner":
        return None
    for label, key in label_to_key.items():
        if len(label) < 2:
            continue
        if label in normalized or normalized in label:
            return key
    if slug == "dynamica":
        return _match_dynamica_grade(header)
    return None


def _parse_header_row(
    headers: list[str], label_to_key: dict[str, str], *, slug: Optional[str] = None
) -> tuple[Optional[int], dict[str, int]]:
    model_idx: Optional[int] = None
    grade_indices: dict[str, int] = {}
    for i, header in enumerate(headers):
        if _is_model_header(header):
            model_idx = i
            continue
        grade_key = _match_grade_column(header, label_to_key, slug=slug)
        if grade_key and grade_key not in grade_indices:
            grade_indices[grade_key] = i
    return model_idx, grade_indices


def _find_header_row(
    rows: list[tuple[Any, ...]], label_to_key: dict[str, str], *, slug: Optional[str] = None
) -> tuple[Optional[int], Optional[int], dict[str, int], list[str]]:
    best: tuple[int, int, Optional[int], dict[str, int], list[str]] = (-1, -1, None, {}, [])
    for row_idx, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        headers = [_normalize_header(cell) for cell in row]
        if not any(headers):
            continue
        model_idx, grade_indices = _parse_header_row(headers, label_to_key, slug=slug)
        score = (10 if model_idx is not None else 0) + len(grade_indices)
        if score > best[0] or (score == best[0] and len(grade_indices) > len(best[3])):
            best = (score, row_idx, model_idx, grade_indices, headers)
    if best[2] is None:
        return None, None, {}, best[4] if best[4] else []
    return best[1], best[2], best[3], best[4]


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color.replace("#", ""), end_color=hex_color.replace("#", ""), fill_type="solid")


def export_company_excel(slug: str) -> bytes:
    company = get_company_by_slug(slug)
    if not company:
        raise ValueError(f"Company not found: {slug}")
    grades = grade_columns_list(company)
    rows = fetch_all(
        """
        SELECT normalized_name, grade, price
        FROM company_prices cp
        JOIN companies c ON c.id = cp.company_id
        WHERE c.slug = %s
        ORDER BY normalized_name, grade
        """,
        (slug,),
    )
    pivot: dict[str, dict[str, int]] = {}
    for r in rows:
        pivot.setdefault(r["normalized_name"], {})[r["grade"]] = r["price"]

    wb = Workbook()
    ws = wb.active
    ws.title = company["name"][:31]
    ws.sheet_view.rightToLeft = True

    headers = ["דגם"] + [g["label"] for g in grades]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for device in sorted(pivot.keys()):
        row = [device] + [pivot[device].get(g["key"]) for g in grades]
        ws.append(row)

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_summary_excel() -> bytes:
    summary = build_summary()
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.rightToLeft = True

    tier_blocks: list[tuple[int, list]] = []
    for tc in summary["tier_config"]:
        cols = []
        for slug in tc["companies"]:
            comp = next(c for c in summary["companies"] if c["slug"] == slug)
            for g in comp["grades"]:
                if g.get("summary_tier") == tc["tier"]:
                    cols.append((slug, comp["name"], g))
        tier_blocks.append((tc["tier"], cols))

    headers = ["דגם"]
    for _, cols in tier_blocks:
        for _, cname, g in cols:
            headers.append(f"{cname} — {g['label']}")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for device in summary["devices"]:
        price_map: dict[tuple[str, str], int] = {}
        for tier_data in device["tiers"]:
            for p in tier_data["prices"]:
                price_map[(p["company"], p["grade_key"])] = p["price"]
        row = [device["normalized_name"]]
        for _, cols in tier_blocks:
            for slug, _, g in cols:
                row.append(price_map.get((slug, g["key"])))
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_company_excel(file_bytes: bytes, slug: str) -> tuple[list[dict], list[str]]:
    company = get_company_by_slug(slug)
    if not company:
        raise ValueError(f"Company not found: {slug}")
    grades = grade_columns_list(company)
    label_to_key = _build_grade_label_map(slug, grades)

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = _load_aligned_rows(ws)
    if not all_rows:
        return [], ["Empty spreadsheet"]

    header_row_idx, model_idx, grade_indices, found_headers = _find_header_row(
        all_rows, label_to_key, slug=slug
    )
    if model_idx is None:
        visible = [h for h in found_headers if h]
        hint = f" Found headers: {visible}" if visible else ""
        return [], [
            "Could not find model column. Expected one of: דגם, Model, Device."
            + hint
        ]
    if not grade_indices:
        return [], ["Could not find any grade price columns for this company"]

    leading_blanks = _header_leading_blanks(found_headers)
    grade_col_idx = next(iter(grade_indices.values()))

    parsed: list[dict] = []
    errors: list[str] = []
    data_start_row = header_row_idx + 2 if header_row_idx is not None else 2
    for offset, row in enumerate(all_rows[header_row_idx + 1 :], start=data_start_row):
        if not row or all(c is None for c in row):
            continue
        model_col, grade_col = _resolve_row_columns(row, model_idx, grade_col_idx, leading_blanks)
        raw_name = str(row[model_col]).strip() if model_col < len(row) and row[model_col] else ""
        if not raw_name:
            continue
        norm = normalize_device_name(raw_name)
        if not norm:
            errors.append(f"Row {offset}: could not normalize '{raw_name}'")
            continue
        for grade_key, col_idx in grade_indices.items():
            _, grade_col = _resolve_row_columns(row, model_idx, col_idx, leading_blanks)
            val = row[grade_col] if grade_col < len(row) else None
            price = _parse_price(val)
            if price is None:
                if val is not None and str(val).strip():
                    errors.append(f"Row {offset}, grade {grade_key}: invalid price '{val}'")
                continue
            parsed.append(
                {
                    "raw_device_name": raw_name,
                    "normalized_name": norm.normalized_name,
                    "brand": norm.brand,
                    "model": norm.model,
                    "storage_gb": norm.storage_gb,
                    "grade": grade_key,
                    "price": price,
                }
            )
    parsed, dedupe_warnings = dedupe_price_rows(parsed)
    errors.extend(dedupe_warnings)
    return parsed, errors


def preview_import(file_bytes: bytes, slug: str, limit: int = 10) -> dict[str, Any]:
    rows, errors = parse_company_excel(file_bytes, slug)
    detected: dict[str, str] = {}
    company = get_company_by_slug(slug)
    if company:
        label_to_key = _build_grade_label_map(slug, grade_columns_list(company))
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        all_rows = _load_aligned_rows(wb.active)
        _, _, grade_indices, found_headers = _find_header_row(all_rows, label_to_key, slug=slug)
        detected = {grade: found_headers[idx] for grade, idx in sorted(grade_indices.items())}
    return {
        "row_count": len(rows),
        "preview": rows[:limit],
        "errors": errors[:20],
        "error_count": len(errors),
        "detected_columns": detected,
    }
